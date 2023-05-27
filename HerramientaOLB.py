###################################################################################################################
######Creado por: Andrey Astorga Bogantes
######Ultima Actualización:07/05/2023
######Parte del PFG para optar por el grado de licenciatura en Ing.Mecatrónica
######Area Academica de Ingenieria en Mecatronica, ITCR.
###################################################################################################################

#Se importan las librerias necesarias para el dessarrollo
import os
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from scipy.stats import norm
import argparse
import imutils
import cv2
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox

#Variable Globales, contiene los valores de GV promedios, agudeza y el nombre del archivo de excel generado
dataGV=[]
dataAcut=[]
ArchivoExcel="GV and Acutance.xlsx"

#Entradas: recibe el directio (path) de la imagen analizar
#Salidas: imagen en formato png 
#Funcion: carga la imagen indicada en escala de grises,genera una imagen del histograma y la curva tonal de
#la imagen analizada

def histograma(pathImagen):
    
    #path Predeterminado para hacer pruebas
    #pathImagen=r"C:\Users\Andrey\Desktop\PFG proyecto\1.jpg"

    #Saca el directorio de la ubicacion de la imagen usango la libreria os
    directorio = os.path.dirname(pathImagen)

    # Utiliza la libreria CV2 para abrir la imagen en una carpeta dada, en su formato original
    imageProyectoN = cv2.imread(pathImagen)
    #Se convierte la imagen a escala de grises
    imageProyecto = cv2.cvtColor(imageProyectoN, cv2.COLOR_BGR2RGB)
    
    #Parte de la verificación del proceso realizado, no es necesario
    #cv2.imwrite('EscalaGrises-Imagen.png', imageProyecto)

    # Crea el histograma de la imagen, utiliza la libreria de numpy
    # Se elige utilizar los niveles de 0 hasta 255, por lo que se tinene 256 niveles
    histogram, bin_edges = np.histogram(imageProyecto, bins=256,)


    # Titulo y nombre de los ejes del histograma
    plt.figure()
    plt.title("Histograma")
    plt.xlabel("Valor en la escala de grises")
    plt.ylabel("Cantidad de pixeles")
      
    #Se extrae el nombre de la imagen que se analiza
    NameUnit = pathImagen.split("\\")[-1].split(".")[0]
    #Se crea un nuevo nombre de archivo en png
    nombreArchivo="Histograma unidad "+NameUnit+".png"
    #Se crea una nueva carpeta para guardar las imagenes
    carpetaPath=os.path.join(directorio,"Histogramas")
    #Directorio para  la imagen
    pathHistograma = os.path.join(carpetaPath, nombreArchivo)
    #Se verifica la existencia de la carpeta, de lo contrario se crea
    if not os.path.exists(carpetaPath):
        os.makedirs(carpetaPath)
    #Plotea los valores del histograma
    plt.plot(histogram)
    #Guarda el histograma, Debido a que es un grafico generado, se tiene que guardar con la libreria matplotlib
    plt.savefig(pathHistograma, bbox_inches='tight')
    plt.close()
    


    #Para verficacion
    #plt.show()
    return


#Entradas: recibe el directio (path) de la imagen analizar
#Salidas: datos de gris promedios de la imagen
#Funcion: analizar el valor GV promedio de las imagenes indicadas. Guarda los datos en una hoja de excel indicando la
#ubicacion de la imagen y el resultado obtenido

def grisPromedio(pathImagen):
    #Se carga la imagen en escala de grises
    image = cv2.imread(pathImagen, cv2.IMREAD_GRAYSCALE)

    #Variables totales, contienen el valor de GV y la cantidad total de pixeles
    total_GV= 0
    total_Pixels = 0

    #Se itera para leer cada pixel de la imagen mediante un doble loop
    #se actualizan las variables
    for row in image:
        for pixel in row:
            total_GV += pixel
            total_Pixels += 1

    # Por definicion se obtiene el valor gris promedio como el valor de GV entre el total de pixeles
    medianGV = total_GV / total_Pixels

    #se guarda los resultados obtenidos en una lista
    dataGV.append({"File Path":pathImagen, "Median Gray Value": medianGV})
    # se crea una variable de data frame (df) en este caso de los resultados de GV de la lista
    df = pd.DataFrame(dataGV)
    # se guardan los datos en una hoja de excel
    df.to_excel(ArchivoExcel,"GV Mean", index=False)
    return


#Entradas: recibe el directio (path) de la imagen analizar
#Salidas: imagen en formato png 
#Funcion: realiza una umbralizacion adaptativa de la imagen ,despues guarda el archivo

def saturacion(pathImagen):
    #Se carga la imagen en escala de grises
    img = cv2.imread(pathImagen, cv2.IMREAD_GRAYSCALE)
    

    #Otros metodos de umbralizacion, no se utilizan pero se pueden usar como referencia
    #_, satur_Bin = cv2.threshold(img, 0, 255, cv2.THRESH_BINARY+cv2.THRESH_OTSU)
    #th, satur_Bin = cv2.threshold(img, 215, 255, cv2.THRESH_TOZERO_INV)

    #Funcion de la libreria de CV2 para realizar umbralizacion adaptativa
    satur_Bin = cv2.adaptiveThreshold(img, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)
    
    #Saca el directorio de la ubicacion de la imagen usango la libreria os
    directorio = os.path.dirname(pathImagen)
    #Se extrae el nombre de la imagen que se analiza
    NameUnit = pathImagen.split("\\")[-1].split(".")[0]
    #Nuevo nombre del archivo
    nombreArchivo="Umbral Adaptativo Unidad; "+NameUnit+".png"
    #nUevo nombre de la carpeta
    carpetaPath=os.path.join(directorio,"Analisis Umbral")
    #Nuevo directorio de la imagen a guardar
    pathUmbral= os.path.join(carpetaPath, nombreArchivo)
    #se verifica la existencia de la carpeta, sino se crea
    if not os.path.exists(carpetaPath):
        os.makedirs(carpetaPath)
    #Se guarda la imagen
    cv2.imwrite(pathUmbral, satur_Bin)
    return 

#Entradas: recibe el directio (path) de la imagen analizar
#Salidas: imagen en formato png 
#Funcion: identifica en la imagen las zonas donde se presenta la mayor cantidad de luz

def zonasBrillo(pathImagen):
    #Se carga la imagen en formato original y luego se convierte en escala de grises
    img = cv2.imread(pathImagen)
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    #Se calcula el brillo promedio utilizando numpy
    BrilloProm = np.mean(gray)

    #Se define un nivel de analisis, como el 50% de la luz es mayor
    #Este nivel sirve para realiza una umbralizacion en la imagen e identificar las zonas
    threshold = BrilloProm * 1.5
    _, thresh = cv2.threshold(gray, threshold, 255, cv2.THRESH_BINARY)

    # Una vez identificados se buscan los contornos para dibujar los rectangulos en las
    #zonas de interes, medianto un for
    contours, _ = cv2.findContours(thresh, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
    for contour in contours:
        (x, y, w, h) = cv2.boundingRect(contour)
        cv2.rectangle(img, (x, y), (x + w, y + h), (0, 0, 255), 2)

   
    #Saca el directorio de la ubicacion de la imagen usango la libreria os
    directorio = os.path.dirname(pathImagen)
    #Se saca el nombre de la unidad a analizar
    NameUnit = pathImagen.split("\\")[-1].split(".")[0]
    #Nuevo nombre del archivo que se va guardar en png
    nombreArchivo="Zonas de brillo Unidad; "+NameUnit+".png"
    #Ubicacion de la carpeta
    carpetaPath=os.path.join(directorio,"Analisis Zonas de Brillo")
    #Ubicacion de la imagen
    pathImagenBrillo= os.path.join(carpetaPath, nombreArchivo)
    #Verificacion de la carpeta, si no existe la crea
    if not os.path.exists(carpetaPath):
        os.makedirs(carpetaPath)
    #Guarda el archivo
    cv2.imwrite(pathImagenBrillo, img)
    
    return


#Entradas: recibe el directio (path) de la imagen analizar
#Salidas: datos de agudeza de la imagen
#Funcion: analizar el valor de agudeza de las imagenes indicadas. Guarda los datos en una hoja de excel indicando la
#ubicacion de la imagen y el resultado obtenido.

def agudeza(pathImagen):
    #lee la imagen en escala de grises
    img = cv2.imread(pathImagen,0)

    #Manipulacion matematica para obtener la agudeza, mediante laPLace, promedios y raices
    laplacian = cv2.Laplacian(img,cv2.CV_64F)
    gnorm = np.sqrt(laplacian**2)
    sharpness = np.average(gnorm)

    #Se guarda la data generada en una lista
    dataAcut.append({"File Path": pathImagen, "Acutance": sharpness})

    #La informacion de la lista se transforma en una data frame, para guardarlo en excel
    df2 = pd.DataFrame(dataAcut)
    #Para tener un solo archivo de excel, se utiliza el siguiente metodo, el cual trata de acceder a la hoja especificada,
    #si no la encuentra la crea dentro del archivo de excel. Recordar que el excel tambien contiene el GV medio
    with pd.ExcelWriter(ArchivoExcel, engine="openpyxl", mode="a") as writer:
        df2.to_excel(writer, "Acutance", index=False)
 
    return

#Entradas:ninguno
#Salidas: actualiza una variable
#Funcion: despliega un cuadro de seleccion para el usuario, para que este escoja la carpeta con las imagenes a analizar
def buscarCarpeta():
    #Pide al usuarion que selecicone la carpeta
    folder_path = filedialog.askdirectory()
    #antes de guardar la informacion, elimina cualquier dato almacenado para la variable
    folder_entry.delete(0, tk.END)
    #guarda el path seleccionado como parte de los widget de tkinter
    folder_entry.insert(0, folder_path)


#Entradas: directorio del archivo de excel que se genera con la herramienta
#Salidas: modificacion de un archvio de excel
#Funcionamiento: del excel que se genera mediante el resto del analisis genera una nueva hoja que funciona como resumen
#de valores estadisticos, como promedio y desviacion, ademas de genera un grafico para ver el comportamiento de la
#agudeza de las imagenes analizadas. Se hace una funcion aparte por temas de indexacion.
    
def estadisticaA(pathExcel):
    #Obtiene la direccion de la carpeta donde se encuentran las imagenes.
    pathDir=os.path.dirname(pathExcel)

    #Lee el excel y lo asigna a un entorno de data frame, tambien lo abre
    #para poder editar el contenido de este. El data frame sirve para darle
    #el formato adecuado a la data generada.
    df = pd.read_excel(pathExcel,sheet_name='Acutance')
    archivoExcel = load_workbook(pathExcel)

    #Se crea una nueva hoja de excel con un nombre dado, en este caso
    #relacionado a la variable de agudeza de la imagen 
    AcutanceSheet = archivoExcel.create_sheet(title='Statistics Acutance')

    # Promedio y desviacion estandar para la columna de datos de agudeza
    #Aqui se utiliza el Data Frame del archivo de excel
    promedioA = df['Acutance'].mean()
    stdDevA = df['Acutance'].std()

    #Tabulacion de datos
    dataT = {'Statistic Acutance': ['Mean', 'Standard Deviation'],
                  'Value': [promedioA, stdDevA]}
    # Se convierte la tabla al formato del data fram
    table_df = pd.DataFrame(dataT)

    #Descompone la tabla en valores, y por medio de un loop guarda
    #los datos
    for row in dataframe_to_rows(table_df, index=False, header=True):
         AcutanceSheet.append(row)

    #Cogido encragado de generar un plot con los datos obtenidos
    fig, ax = plt.subplots()
    #cantidad de valores
    x = np.arange(len(df['Acutance']))
    #Valores
    y = df['Acutance']
    #leyendas
    ax.plot(x, y, label='Data')
    ax.plot(x, np.poly1d(np.polyfit(x, y, 1))(x), label='Trend')
    #Etiquetas
    ax.set_xlabel('Number of image')
    ax.set_ylabel('Value of Acutance')
    ax.set_title('Trend Curve Acutance')
    ax.legend()
    
    #Colocacion de los valores de desviacion y promedio en el grafico.
    #Se colocan al final.
    ax.text(x[-1], promedioA, f"Mean: {promedioA:.2f}", ha='right', va='center')
    ax.text(x[-1], promedioA-stdDevA, f"Std Deviation: {stdDevA:.2f}", ha='right', va='center')

    #Se guarda como imagen en la carpeta donde se hace el analisis
    #Necesita el directorio de la carpeta y un nombre para la imagen
    ImageName='Trend Acutance.png'
    graph_filename = os.path.join(pathDir, ImageName)
    plt.savefig(graph_filename)
    #se cierra plt para ahorrar memoria y evitar errores.
    #cuando se trabaja con mayor cantidad de datos,
    plt.close()

    
    graph_img = Image(graph_filename)
    graph_img.anchor = 'E2'  
    #Agrega la imagen en la posicion o celda indicada.
    AcutanceSheet.add_image(graph_img)

    #Se guarda el excel.
    archivoExcel.save(pathExcel)
    return

#Entradas: directorio del archivo de excel que se genera con la herramienta
#Salidas: modificacion de un archvio de excel
#Funcionamiento: del excel que se genera mediante el resto del analisis genera una nueva hoja de excel que funciona como resumen
#de valores estadisticos, como promedio y desviacion, ademas de genera un grafico para ver el comportamiento del
#valor de gris promedio de las imagenes analizadas. Se hace una funcion aparte por temas de indexacion.

def estadisticaGV(pathExcel):
    #Obtiene la direccion de la carpeta donde se encuentran las imagenes.
    pathDir=os.path.dirname(pathExcel)

    #Lee el excel y lo asigna a un entorno de data frame, tambien lo abre
    #para poder editar el contenido de este. El data frame sirve para darle
    #el formato adecuado a la data generada.
    df = pd.read_excel(pathExcel,sheet_name='GV Mean')
    archivoExcel = load_workbook(pathExcel)

    #Se crea una nueva hoja de excel con un nombre dado, en este caso
    #relacionado a la variable de agudeza de la imagen 
    GVSheet = archivoExcel.create_sheet(title='Statistics GV Mean')

    # Promedio y desviacion estandar para la columna de datos de agudeza
    #Aqui se utiliza el Data Frame del archivo de excel
    promedioGV = df['Median Gray Value'].mean()
    stdDevGV = df['Median Gray Value'].std()

    #Tabulacion de datos
    dataT = {'Statistic GV': ['Mean', 'Standard Deviation'],
                  'Value': [promedioGV, stdDevGV]}
    # Se convierte la tabla al formato del data fram
    table_df = pd.DataFrame(dataT)

    #Descompone la tabla en valores, y por medio de un loop guarda
    #los datos
    for row in dataframe_to_rows(table_df, index=False, header=True):
         GVSheet.append(row)

    #Cogido encragado de generar un plot con los datos obtenidos
    fig, ax = plt.subplots()
    #cantidad de valores
    x = np.arange(len(df['Median Gray Value']))
    #Valores
    y = df['Median Gray Value']
    #leyendas
    ax.plot(x, y, label='Data')
    ax.plot(x, np.poly1d(np.polyfit(x, y, 1))(x), label='Trend')
    #Etiquetas
    ax.set_xlabel('Number of image')
    ax.set_ylabel('Value of GV')
    ax.set_title('Trend Curve GV')
    ax.legend()
    
    #Colocacion de los valores de desviacion y promedio en el grafico.
    #Se colocan al final.
    ax.text(x[-1], promedioGV, f"Mean: {promedioGV:.2f}", ha='right', va='center')
    ax.text(x[-1], promedioGV - stdDevGV, f"Std Deviation: {stdDevGV:.2f}", ha='right', va='center')

    #Se guarda como imagen en la carpeta donde se hace el analisis
    #Necesita el directorio de la carpeta y un nombre para la imagen
    ImageName='Trend GV Mean.png'
    graph_filename = os.path.join(pathDir, ImageName)
    plt.savefig(graph_filename)
    #se cierra plt para ahorrar memoria y evitar errores.
    #cuando se trabaja con mayor cantidad de datos,
    plt.close()

    
    graph_img = Image(graph_filename)
    graph_img.anchor = 'E2'  
    #Agrega la imagen en la posicion o celda indicada.
    GVSheet.add_image(graph_img)

    #Se guarda el excel.
    archivoExcel.save(pathExcel)
    return   


#Entradas: recibe el directorio de la carpeta donde se guardan las imagenes
#Salidas:ninguna
#Funcion: hace el llamado a todas las funciones para analizar las imagenes que interesan. Debe acceder a cada imagen
def main(folder_entry):
    #accede a la variable global que contiene el nombre del archivo de excel
    global ArchivoExcel
    folder_path=folder_entry
    #actualiza la ubicacion donde se va a guardar el archivo de excel
    pathExcel= os.path.join(folder_path, ArchivoExcel)
    #asigna el nuevo valor
    ArchivoExcel= pathExcel

    #Variable para contar la cantidad unidades analizadas
    contadorUnidades=0

    # for para analizar los archivos dentro de la carpeta.
    # como solo interesan los archivos de imagen en jpg, se define solo para este tipo de archivo
    for filename in os.listdir(folder_path):
        #tipo de archivos de interes
        if filename.endswith(".jpg"):
            #actualiza el directorio de la imagen a analizar
            PathImagen = os.path.join(folder_path, filename)
            #actualiza el contador
            contadorUnidades+=1
            #Llama a las funciones
            histograma(PathImagen)
            grisPromedio(PathImagen)
            zonasBrillo(PathImagen)
            saturacion(PathImagen)
            agudeza(PathImagen)
    if contadorUnidades>1:
        estadisticaGV(ArchivoExcel)
        estadisticaA(ArchivoExcel)
    
    import time
    #despliega un mensaje que se ha finalizado el programa, pero espera 2 segundos primero
    cantidadImg="Imagenes analizadas: "+str(contadorUnidades)
    time.sleep(2)
    messagebox.showinfo("Proceso Completado", cantidadImg)
    
#Commandos relacionados a la creacion de una interfaz visual para pedir la información al usuario
#Crea una ventana, con un titulo y de un tamaño dado de 400x200
window = tk.Tk()
window.title("Herramienta para analizar imagenes de OLB")
window.geometry("400x200")

#Crea una etiqueta
folder_label = tk.Label(window, text="Inserte la direccion de la Carpeta con las imagenes:")
folder_label.pack()

# Crea un widget para guardar la ubicación de la carpeta, de tamaño 40
folder_entry = tk.Entry(window, width=40)
folder_entry.pack()

# Se crea un boton que permite explorar las carpetas. utiliza la funcion buscarCarpeta
select_button = tk.Button(window, text="1. Explorar este equipo", command=buscarCarpeta)
select_button.pack()

# Se crea un boton para empezar con el proceso de analisis de la imagen,mediante las diferentes funciones generadas
# El proceso inicia debido a que se llama a al funcion main con el directorio de la carpeta y se llaman las funciones
process_button = tk.Button(window, text="2. Procesar las Imagenes", command=lambda: main(folder_entry.get()))
process_button.pack()

# Loop para el tkinter
window.mainloop()

